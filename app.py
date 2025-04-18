import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import scrolledtext
from tkinter.ttk import Progressbar, Frame, Label, Button, Entry
from ttkbootstrap import Style
import openpyxl
from urllib.parse import quote
import webbrowser
import time
import pyautogui
import logging
from threading import Thread, Event

# Configurar logging
logging.basicConfig(filename='whatsapp_automation.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

stop_event = Event()
logs = {"success": [], "failure": []}


def criar_modelo_planilha():
    # Cria um modelo de planilha com colunas "Nome" e "Telefone"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Modelo"
    sheet.append(["Nome", "Telefone"])
    # Adiciona uma linha de exemplo
    # Número de exemplo com código do país (Brasil)
    sheet.append(["Exemplo Nome", "5513111111112"])
    return workbook


def baixar_modelo():
    # Função para salvar o modelo de planilha no local especificado pelo usuário
    modelo_workbook = criar_modelo_planilha()
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not save_path:
        return

    modelo_workbook.save(save_path)
    messagebox.showinfo("Sucesso", "Modelo de planilha baixado com sucesso!")


def enviar_mensagens(file_path, wait_time_whatsapp, wait_time_message, mensagem_template):
    try:
        # Abrir WhatsApp Web
        webbrowser.open('https://web.whatsapp.com/')
        # Aguarda para que o WhatsApp Web carregue
        time.sleep(wait_time_whatsapp)

        # Ler planilha e guardar informações sobre nome e telefone
        workbook = openpyxl.load_workbook(file_path)
        pagina_demandas = workbook.active

        total_rows = pagina_demandas.max_row - 1
        current_row = 0

        for linha in pagina_demandas.iter_rows(min_row=2, values_only=True):
            if stop_event.is_set():
                break

            Nome, Telefone = linha

            if not Nome or not Telefone:
                continue

            # Formatar telefone: remover caracteres especiais e adicionar o código do país
            Telefone = str(Telefone).replace('(', '').replace(
                ')', '').replace('-', '').replace(' ', '')
            if not Telefone.startswith('55'):
                Telefone = '55' + Telefone

            # Verificar se o telefone está no formato correto do WhatsApp (55 + número)
            if len(Telefone) != 13:
                logging.error(f"Número de telefone inválido para {
                              Nome} ({Telefone})")
                log_text.insert(tk.END, f"Número de telefone inválido para {
                                Nome} ({Telefone})\n")
                logs["failure"].append((Nome, Telefone))
                continue

            # Mensagem personalizada
            mensagem = mensagem_template.replace("{Nome}", Nome)

            # Criar link personalizado do WhatsApp
            link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={
                Telefone}&text={quote(mensagem)}'
            webbrowser.open(link_mensagem_whatsapp)
            time.sleep(wait_time_message)  # Ajuste o tempo conforme necessário

            try:
                # Pressionar Enter para enviar a mensagem
                pyautogui.press('enter')
                # Pausa para garantir que a mensagem seja enviada
                time.sleep(6)

                # Fechar a aba atual
                pyautogui.hotkey('ctrl', 'w')
                time.sleep(2)

                # Registrar sucesso no log
                logging.info(f'Mensagem enviada para {Nome} ({Telefone})')
                log_text.insert(tk.END, f'Mensagem enviada para {
                                Nome} ({Telefone})\n')
                logs["success"].append((Nome, Telefone))
            except Exception as e:
                logging.error(f"Erro ao enviar mensagem para {
                              Nome} ({Telefone}): {str(e)}")
                log_text.insert(tk.END, f"Erro ao enviar mensagem para {
                                Nome} ({Telefone}): {str(e)}\n")
                logs["failure"].append((Nome, Telefone))

            log_text.see(tk.END)

            # Atualizar barra de progresso
            current_row += 1
            progress_var.set((current_row / total_rows) * 100)
            progress_bar.update()

        # Garantir que a barra de progresso chegue a 100% no final
        progress_var.set(100)
        progress_bar.update()
        messagebox.showinfo("Sucesso", "Mensagens enviadas com sucesso!")
    except Exception as e:
        logging.error(f"Erro ao enviar mensagens: {str(e)}")
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")


def iniciar_envio():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        messagebox.showerror("Erro", "Nenhuma planilha selecionada.")
        return

    try:
        wait_time_whatsapp = int(entry_wait_time_whatsapp.get())
        wait_time_message = int(entry_wait_time_message.get())
        mensagem_template = text_mensagem.get("1.0", tk.END).strip()
        if "{Nome}" not in mensagem_template:
            raise ValueError("A mensagem deve conter a variável {Nome}.")
    except ValueError as e:
        messagebox.showerror("Erro", f"Entrada inválida: {str(e)}")
        return

    stop_event.clear()
    thread = Thread(target=enviar_mensagens, args=(
        file_path, wait_time_whatsapp, wait_time_message, mensagem_template))
    thread.start()


def parar_envio():
    stop_event.set()
    messagebox.showinfo(
        "Interrupção", "O envio das mensagens foi interrompido.")


def salvar_logs():
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not save_path:
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Logs de Envio"

    sheet.append(["Status", "Nome", "Telefone"])
    for status, entries in logs.items():
        for nome, telefone in entries:
            sheet.append([status, nome, telefone])

    workbook.save(save_path)
    messagebox.showinfo("Sucesso", "Logs salvos com sucesso!")


def show_info_whatsapp():
    messagebox.showinfo(
        "Informação", "Tempo de espera para carregar WhatsApp: O tempo (em segundos) necessário para que o WhatsApp Web carregue completamente antes de iniciar o envio das mensagens.")


def show_info_message():
    messagebox.showinfo(
        "Informação", "Tempo de espera entre mensagens: O tempo (em segundos) necessário entre o envio de cada mensagem para garantir que o WhatsApp Web processe e envie a mensagem corretamente.")


def show_info_mensagem():
    messagebox.showinfo(
        "Informação", "Mensagem: Insira a mensagem a ser enviada. Utilize {Nome} como um espaço reservado que será substituído pelo nome do destinatário.")


# Configurar a interface gráfica
root = tk.Tk()
root.title("Envio Automático de Mensagens no WhatsApp")

style = Style(theme="lumen")

# Seção de tempos de espera
frame_tempo_espera = Frame(root, padding=10)
frame_tempo_espera.pack(padx=20, pady=10, fill="x", anchor="center")

Label(frame_tempo_espera, text="Tempo de espera para carregar WhatsApp:").grid(
    row=0, column=0, sticky="e")
entry_wait_time_whatsapp = Entry(frame_tempo_espera, width=5)
entry_wait_time_whatsapp.insert(0, "60")
entry_wait_time_whatsapp.grid(row=0, column=1, padx=(5, 10))

Label(frame_tempo_espera, text="segundos").grid(row=0, column=2, sticky="w")
info_button_whatsapp = Button(
    frame_tempo_espera, text="ℹ️", command=show_info_whatsapp)
info_button_whatsapp.grid(row=0, column=3, padx=5)

Label(frame_tempo_espera, text="Tempo de espera entre mensagens:").grid(
    row=1, column=0, sticky="e")
entry_wait_time_message = Entry(frame_tempo_espera, width=5)
entry_wait_time_message.insert(0, "20")
entry_wait_time_message.grid(row=1, column=1, padx=(5, 10))

Label(frame_tempo_espera, text="segundos").grid(row=1, column=2, sticky="w")
info_button_message = Button(
    frame_tempo_espera, text="ℹ️", command=show_info_message)
info_button_message.grid(row=1, column=3, padx=5)

# Seção de mensagem
frame_mensagem = Frame(root, padding=10)
frame_mensagem.pack(padx=20, pady=10, fill="x", anchor="center")

Label(frame_mensagem, text="Mensagem:").grid(row=0, column=0, sticky="w")
text_mensagem = scrolledtext.ScrolledText(frame_mensagem, height=5, width=40)
text_mensagem.insert(
    tk.END, "Olá {Nome}, conto com seu voto! https://www.instagram.com/dadoseleitorais")
text_mensagem.grid(row=1, column=0, columnspan=3, pady=10, padx=(0, 5))

info_button_mensagem = Button(
    frame_mensagem, text="ℹ️", command=show_info_mensagem)
info_button_mensagem.grid(row=1, column=3, padx=5, sticky="n")


# Botões de controle
frame_controle = Frame(root, padding=10)
frame_controle.pack(padx=20, pady=10, fill="x")

# Primeira coluna de botões
botao_baixar_modelo = Button(
    frame_controle, text="Baixar Modelo de Planilha", command=baixar_modelo)
botao_baixar_modelo.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

botao_abrir = Button(
    frame_controle, text="Carregar Planilha e Enviar Mensagens", command=iniciar_envio)
botao_abrir.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

# Segunda coluna de botões
botao_parar = Button(frame_controle, text="Parar Envio", command=parar_envio)
botao_parar.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

botao_salvar_logs = Button(
    frame_controle, text="Salvar Logs", command=salvar_logs)
botao_salvar_logs.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

# Ajuste para expandir as colunas igualmente
frame_controle.columnconfigure(0, weight=1)
frame_controle.columnconfigure(1, weight=1)


# Barra de progresso
frame_progresso = Frame(root, padding=10)
frame_progresso.pack(padx=20, pady=10, fill="x", anchor="center")

progress_var = tk.DoubleVar()
progress_bar = Progressbar(frame_progresso, variable=progress_var, maximum=100)
progress_bar.pack(fill="x", padx=10, pady=10)

# Área de log
frame_log = Frame(root, padding=10)
frame_log.pack(padx=20, pady=10, fill="both", expand=True)

Label(frame_log, text="Logs:").pack(anchor="w", padx=10)
log_text = scrolledtext.ScrolledText(frame_log, height=10)
log_text.pack(fill="both", expand=True, padx=10, pady=10)

# Iniciar a interface
root.mainloop()

